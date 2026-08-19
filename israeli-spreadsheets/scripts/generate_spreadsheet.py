#!/usr/bin/env python3
"""Generate Israeli financial Excel spreadsheets.

Creates formatted Hebrew spreadsheets with Israeli tax calculations,
VAT, NIS formatting, and common financial templates.

Usage:
    python generate_spreadsheet.py --template invoice --output invoice.xlsx
    python generate_spreadsheet.py --template salary --output salary.xlsx
    python generate_spreadsheet.py --template arnona --output arnona.xlsx
    python generate_spreadsheet.py --help

Requirements:
    pip install openpyxl
"""

import argparse
import sys
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Israeli financial constants (2026, VAT at 18% since 2025-01-01)
# Income tax brackets reflect the 2026 bracket widening (chapter C of the Economic Efficiency Law 2026, published 31.3.2026) (effective 1 January 2026):
# the 20% and 31% bands were widened, pushing the 35% floor from 269,280
# to 301,200 NIS annually. The 10/14/35/47 rate thresholds are otherwise
# frozen at 2025 levels through 2027.
VAT_RATE = Decimal("0.18")
TAX_BRACKETS = [
    (Decimal("84120"), Decimal("0.10")),
    (Decimal("120720"), Decimal("0.14")),
    (Decimal("228000"), Decimal("0.20")),
    (Decimal("301200"), Decimal("0.31")),
    (Decimal("560280"), Decimal("0.35")),
    (Decimal("999999999"), Decimal("0.47")),
]
# The statutory ladder ends at 47%. The often-quoted "50%" is 47% plus the
# Section 121B(a) surtax of 3% on taxable income above SURTAX_THRESHOLD, and it
# applies only to personal-exertion income. Section 121B(a1) adds a further 2%
# on capital-source income above the same threshold.
SURTAX_THRESHOLD = Decimal("721560")
SURTAX_RATE = Decimal("0.03")
SURTAX_CAPITAL_EXTRA_RATE = Decimal("0.02")
CREDIT_POINT_VALUE = Decimal("2904")
BL_REDUCED_STEP = Decimal("7703")      # monthly reduced-collection step, 2026
BL_MAX_INSURABLE = Decimal("51910")    # monthly maximum insurable income, 2026
PENSION_EMPLOYEE_RATE = Decimal("0.06")
KH_EMPLOYEE_RATE = Decimal("0.025")
SEC_45A_CREDIT_RATE = Decimal("0.35")  # Section 45A credit on the employee deposit
RESIDENT_CREDIT_POINTS = Decimal("2.25")  # male resident; a female resident gets 2.75
FEMALE_RESIDENT_CREDIT_POINTS = Decimal("2.75")

NIS_FORMAT = '#,##0.00 "₪"'
PERCENT_FORMAT = "0.00%"


def setup_rtl_sheet(ws, title):
    """Configure worksheet for RTL Hebrew display."""
    ws.title = title
    ws.sheet_view.rightToLeft = True


def style_header(ws, row, cols, fill_color="1F4E79"):
    """Style header row with colors."""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Heebo", size=11, bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_invoice(output_path):
    """Create a Hebrew tax invoice template."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "חשבונית מס")

    # Business details
    ws["A1"] = "חשבונית מס / קבלה"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)
    ws["A2"] = "שם העסק: [שם]"
    ws["A3"] = "ע.מ./ח.פ.: [מספר]"
    ws["A4"] = "כתובת: [כתובת]"
    ws["A5"] = "טלפון: [טלפון]"

    # Invoice details
    ws["D2"] = "מספר חשבונית:"
    ws["E2"] = "[מספר]"
    ws["D3"] = "תאריך:"
    ws["E3"] = "[DD/MM/YYYY]"

    # Item headers
    headers = ["תיאור", "כמות", "מחיר ליחידה", "סה\"כ"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=7, column=i, value=h)
    style_header(ws, 7, 4)

    # Sample rows: line total = quantity * unit price
    for row in range(8, 11):
        ws.cell(row=row, column=1, value="[פריט]")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=0).number_format = NIS_FORMAT
        ws.cell(row=row, column=4, value=f"=B{row}*C{row}").number_format = NIS_FORMAT

    # Totals. VAT_RATE is the single source of truth for the rate.
    vat_pct = int(VAT_RATE * 100)
    ws.cell(row=12, column=3, value="סכום ביניים:").font = Font(bold=True)
    ws.cell(row=12, column=4, value="=SUM(D8:D10)").number_format = NIS_FORMAT
    ws.cell(row=13, column=3, value=f"מע\"מ ({vat_pct}%):").font = Font(bold=True)
    ws.cell(row=13, column=4, value=f"=ROUND(D12*{VAT_RATE},2)").number_format = NIS_FORMAT
    ws.cell(row=14, column=3, value="סה\"כ לתשלום:").font = Font(name="Heebo", bold=True, size=12)
    ws.cell(row=14, column=4, value="=D12+D13").number_format = NIS_FORMAT

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    wb.save(output_path)
    print(f"Created invoice template: {output_path}")


def calculate_income_tax(annual_income, credit_points=None, capital_source=False):
    """Israeli progressive income tax on personal-exertion income.

    credit_points defaults to the male-resident entitlement (2.25). Pass
    FEMALE_RESIDENT_CREDIT_POINTS (2.75) for a female resident, or any other
    total once children, olim or Section 39B miluim points are counted.

    Set capital_source=True to add the Section 121B(a1) limb. Note that the
    bracket ladder itself differs for non-exertion income (31/35/47 with no
    10/14/20 bands), so this function is not a complete model for that case.
    """
    if credit_points is None:
        credit_points = RESIDENT_CREDIT_POINTS

    tax = Decimal("0")
    prev_limit = Decimal("0")
    for limit, rate in TAX_BRACKETS:
        if annual_income <= prev_limit:
            break
        taxable = min(annual_income, limit) - prev_limit
        tax += taxable * rate
        prev_limit = limit

    # Section 121B surtax, computed separately from the bracket ladder.
    above = max(Decimal("0"), annual_income - SURTAX_THRESHOLD)
    surtax_rate = SURTAX_RATE + (SURTAX_CAPITAL_EXTRA_RATE if capital_source
                                 else Decimal("0"))
    tax += above * surtax_rate

    credit = credit_points * CREDIT_POINT_VALUE
    return max(Decimal("0"), tax - credit)


def create_salary_slip(output_path):
    """Create a Hebrew salary slip (tlush maskoret) template."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "תלוש משכורת")

    ws["A1"] = "תלוש משכורת"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)

    ws["A3"] = "שם העובד:"
    ws["B3"] = "[שם]"
    ws["A4"] = "ת.ז.:"
    ws["B4"] = "[מספר]"
    ws["A5"] = "חודש:"
    ws["B5"] = "[חודש/שנה]"

    # Earnings
    ws["A7"] = "תשלומים"
    ws["A7"].font = Font(bold=True)
    style_header(ws, 7, 2, "2E7D32")

    earnings = ["שכר בסיס", "שעות נוספות", "בונוס"]
    for i, label in enumerate(earnings, 8):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=0).number_format = NIS_FORMAT
    ws.cell(row=11, column=1, value="סה\"כ ברוטו").font = Font(bold=True)
    ws.cell(row=11, column=2, value="=SUM(B8:B10)").number_format = NIS_FORMAT

    # Deductions
    ws["A12"] = "ניכויים"
    ws["A12"].font = Font(bold=True)
    style_header(ws, 12, 2, "C62828")

    # Inputs the user must set before the sheet is meaningful.
    ws.cell(row=22, column=1, value="נקודות זיכוי (2.25 לגבר תושב, 2.75 לאישה תושבת)")
    ws.cell(row=22, column=2, value=float(RESIDENT_CREDIT_POINTS))
    ws.cell(row=23, column=1, value="ערך נקודת זיכוי לחודש")
    ws.cell(row=23, column=2, value=float(CREDIT_POINT_VALUE / 12)).number_format = NIS_FORMAT
    ws.cell(row=24, column=1, value="שיעור ביטוח לאומי + בריאות, מדרגה מופחתת")
    ws.cell(row=24, column=2, value=0.0427).number_format = PERCENT_FORMAT
    ws.cell(row=25, column=1, value="שיעור ביטוח לאומי + בריאות, מדרגה מלאה")
    ws.cell(row=25, column=2, value=0.1217).number_format = PERCENT_FORMAT
    ws.cell(row=26, column=1, value=("שנו את שיעורי ב\"ל לפי קטגוריית העובד. "
                                     "לקטין ולפנסיונר עובד השיעור הוא 0."))
    ws.cell(row=26, column=1).font = Font(name="Heebo", size=10, italic=True)
    ws.cell(row=27, column=1, value="מדרגת גביה מופחתת לחודש")
    ws.cell(row=27, column=2, value=float(BL_REDUCED_STEP)).number_format = NIS_FORMAT
    ws.cell(row=28, column=1, value="הכנסה מרבית מבוטחת לחודש")
    ws.cell(row=28, column=2, value=float(BL_MAX_INSURABLE)).number_format = NIS_FORMAT
    ws.cell(row=29, column=1, value="שיעור פנסיה עובד")
    ws.cell(row=29, column=2, value=float(PENSION_EMPLOYEE_RATE)).number_format = PERCENT_FORMAT
    ws.cell(row=30, column=1, value="שיעור קרן השתלמות עובד")
    ws.cell(row=30, column=2, value=float(KH_EMPLOYEE_RATE)).number_format = PERCENT_FORMAT
    ws.cell(row=31, column=1, value="שיעור זיכוי 45א על הפקדת העובד לפנסיה")
    ws.cell(row=31, column=2, value=float(SEC_45A_CREDIT_RATE)).number_format = PERCENT_FORMAT

    # Monthly bracket ladder (personal-exertion income), from the ITA booklet.
    ws.cell(row=13, column=1, value="מס הכנסה (אחרי זיכוי 45א)")
    ws.cell(row=13, column=2, value=(
        "=MAX(0,"
        "MIN(B11,7010)*0.10"
        "+MAX(0,MIN(B11,10060)-7010)*0.14"
        "+MAX(0,MIN(B11,19000)-10060)*0.20"
        "+MAX(0,MIN(B11,25100)-19000)*0.31"
        "+MAX(0,MIN(B11,46690)-25100)*0.35"
        "+MAX(0,B11-46690)*0.47"
        "+MAX(0,B11-60130)*0.03"
        "-B22*B23-B16*B31)"
    )).number_format = NIS_FORMAT

    ws.cell(row=14, column=1, value="ביטוח לאומי + מס בריאות")
    ws.cell(row=14, column=2, value=(
        "=MIN(B11,B27)*B24+MAX(0,MIN(B11,B28)-B27)*B25"
    )).number_format = NIS_FORMAT

    ws.cell(row=15, column=1, value="(מוצג יחד עם ביטוח לאומי)")
    ws.cell(row=15, column=1).font = Font(name="Heebo", size=10, italic=True)

    ws.cell(row=16, column=1, value="פנסיה עובד")
    ws.cell(row=16, column=2, value="=B11*B29").number_format = NIS_FORMAT
    ws.cell(row=17, column=1, value="קרן השתלמות עובד")
    ws.cell(row=17, column=2, value="=B11*B30").number_format = NIS_FORMAT
    ws.cell(row=18, column=1, value="סה\"כ ניכויים").font = Font(bold=True)
    ws.cell(row=18, column=2, value="=B13+B14+B16+B17").number_format = NIS_FORMAT

    # Net
    ws.cell(row=19, column=1, value="שכר נטו").font = Font(name="Heebo", bold=True, size=14)
    ws.cell(row=19, column=2, value="=B11-B18").number_format = NIS_FORMAT

    ws.cell(row=33, column=1, value=(
        "המדרגות כאן הן חודשיות ולהכנסה מיגיעה אישית. שורה 13 מפחיתה את זיכוי "
        "נקודות הזיכוי ואת זיכוי 45א על הפקדת העובד לפנסיה. יש לאמת מול תלוש "
        "רשמי לפני שימוש בפועל."))
    ws.cell(row=33, column=1).font = Font(name="Heebo", size=10, italic=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"Created salary slip: {output_path}")


def create_arnona(output_path):
    """Create an arnona calculator spreadsheet."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "מחשבון ארנונה")

    ws["A1"] = "מחשבון ארנונה"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)

    ws["A2"] = ("הזינו את התעריף השנתי למ\"ר מתוך צו הארנונה של הרשות המקומית "
                "לשנת המס. התעריף משתנה לפי אזור, סוג בניין וגודל הנכס, ולכן "
                "אין תעריף אחד לעיר.")
    ws["A2"].font = Font(name="Heebo", size=10, italic=True)

    headers = ["רשות מקומית", "אזור / סוג בניין", "תעריף שנתי למ\"ר",
               "שטח (מ\"ר)", "ארנונה שנתית", "ארנונה דו-חודשית"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 6)

    # Deliberately blank: municipal tariffs are published per sqm per YEAR in
    # each authority's tzav arnona, stratified by zone, building type and area
    # band. A single hardcoded number per city would be wrong for most
    # properties, so the user fills in the tariff for their own property.
    for row in range(5, 12):
        ws.cell(row=row, column=1, value="[רשות]")
        ws.cell(row=row, column=2, value="[אזור / סוג]")
        ws.cell(row=row, column=3, value=0).number_format = NIS_FORMAT
        ws.cell(row=row, column=4, value=80)  # default 80 sqm
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}").number_format = NIS_FORMAT
        ws.cell(row=row, column=6, value=f"=E{row}/6").number_format = NIS_FORMAT

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 22

    wb.save(output_path)
    print(f"Created arnona calculator: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Israeli financial Excel spreadsheets"
    )
    parser.add_argument(
        "--template", choices=["invoice", "salary", "arnona"],
        default="invoice",
        help="Spreadsheet template (default: invoice)"
    )
    parser.add_argument(
        "--output", default="spreadsheet.xlsx",
        help="Output file path (default: spreadsheet.xlsx)"
    )
    args = parser.parse_args()

    generators = {
        "invoice": create_invoice,
        "salary": create_salary_slip,
        "arnona": create_arnona,
    }
    generators[args.template](args.output)


if __name__ == "__main__":
    main()
